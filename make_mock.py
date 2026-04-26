from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, RowBreak
from openpyxl.worksheet.properties import PageSetupProperties

def EMU(px): return int(px * 9525)

def add_logo_centered(ws, row_height_px, total_width_px, logo_path, logo_h_px=44):
    ratio = 650 / 120
    logo_w_px = int(logo_h_px * ratio)
    x_px = (total_width_px - logo_w_px) / 2
    y_px = (row_height_px - logo_h_px) / 2
    img = XLImage(logo_path)
    img.height = logo_h_px
    img.width  = logo_w_px
    img.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(x=EMU(x_px), y=EMU(y_px)),
        ext=XDRPositiveSize2D(cx=EMU(logo_w_px), cy=EMU(logo_h_px))
    )
    ws.add_image(img)

wb = Workbook()

GREEN  = "FBC10B"   # STARTTOP アクセントイエロー
DGREEN = "051C3D"   # STARTTOP プライマリネイビー
LGRAY  = "F4F4F4"
YELLOW = "FFFBEA"
WHITE  = "FFFFFF"
LGREEN = "FFECAF"   # STARTTOP 薄黄色
RED    = "C00000"
LRED   = "FFEAEA"
LOGO   = "/Users/iwa/claude/solar-simulator/public/starttop_logo.png"

thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="C00000")
BORDER  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
RBORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

def sc(ws, row, col, value="", bg=None, bold=False, align="left",
       fmt=None, fg="222222", size=10, merge_to=None, height=None, wrap=False, bdr=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="メイリオ", size=size, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            indent=(1 if align == "left" else 0), wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    c.border = bdr if bdr else BORDER
    if merge_to:
        ws.merge_cells(f"{get_column_letter(col)}{row}:{merge_to}{row}")
    if height:
        ws.row_dimensions[row].height = height
    return c

def fill_row(ws, row, bg, n_cols=7):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = BORDER

# =================================================================
# Sheet1: 入力シート
# =================================================================
ws1 = wb.active
ws1.title = "入力"

for col, w in zip("ABCD", [24, 16, 10, 20]):
    ws1.column_dimensions[col].width = w

sc(ws1, 1, 1, "", bg=WHITE, merge_to="D", height=56)
add_logo_centered(ws1, row_height_px=56, total_width_px=490, logo_path=LOGO)

sc(ws1, 2, 1, "太陽光発電シミュレーション｜入力シート",
   bg=GREEN, bold=True, fg=DGREEN, size=13, align="center", merge_to="D", height=28)
sc(ws1, 3, 1, "黄色のセルに数値を入力してください",
   bg=DGREEN, bold=True, fg=WHITE, merge_to="D", height=22)
sc(ws1, 4, 1, "項目",   bg=GREEN, bold=True, fg=DGREEN, align="center", height=20)
sc(ws1, 4, 2, "入力値", bg=GREEN, bold=True, fg=DGREEN, align="center")
sc(ws1, 4, 3, "単位",   bg=GREEN, bold=True, fg=DGREEN, align="center")
sc(ws1, 4, 4, "備考",   bg=GREEN, bold=True, fg=DGREEN, align="center")

inputs = [
    ("お客様名",         "増田",  "",       None,    ""),
    ("パネル容量",       7.92,    "kW",     "0.00",  ""),
    ("蓄電池容量",       9.5,     "kWh",    "0.0",   ""),
    ("お客様電気代",     14000,   "円/月",  "#,##0", ""),
    ("太陽光自家消費量", 100,     "kWh/月", "#,##0", "50〜150kWhで設定"),
    ("年間発電量",       "=B6*1100", "kWh/年", "#,##0", "※パネル容量×1100（手修正可）"),
]
for i, (label, val, unit, fmt, note) in enumerate(inputs):
    row = 5 + i
    ws1.row_dimensions[row].height = 24
    sc(ws1, row, 1, label, bg=LGRAY)
    sc(ws1, row, 2, val,   bg=YELLOW, bold=True, align="right", fmt=fmt)
    sc(ws1, row, 3, unit,  bg=LGRAY, align="center")
    sc(ws1, row, 4, note,  bg=LGRAY, fg="888888")

ws1.row_dimensions[12].height = 8
sc(ws1, 13, 1, "固定値（変更不要）", bg=GREEN, bold=True, fg=DGREEN, merge_to="D", height=20)
fixed = [
    ("買電単価",              30,   "円/kWh", ""),
    ("売電単価 年1〜4年目",   24,   "円/kWh", "FIT"),
    ("売電単価 年5〜10年目",  8.3,  "円/kWh", ""),
    ("売電単価 年11〜30年目", 7,    "円/kWh", "卒FIT"),
    ("基本料金（最低電気代）", 1500, "円/月",  "電気代の下限"),
]
for i, (label, val, unit, note) in enumerate(fixed):
    row = 14 + i
    ws1.row_dimensions[row].height = 20
    sc(ws1, row, 1, label, bg=LGRAY, fg="555555")
    sc(ws1, row, 2, val,   bg=LGRAY, bold=True, align="right", fmt="0.0", fg="555555")
    sc(ws1, row, 3, unit,  bg=LGRAY, align="center", fg="555555")
    sc(ws1, row, 4, note,  bg=LGRAY, fg="888888")

sc(ws1, 19, 1, "参考（自動計算）", bg=GREEN, bold=True, fg=DGREEN, merge_to="D", height=20)

ws1.row_dimensions[20].height = 20
sc(ws1, 20, 1, "月間発電量（参考）", bg=LGRAY, fg="555555")
sc(ws1, 20, 2, "=B10/12",           bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 20, 3, "kWh/月",            bg=LGRAY, align="center", fg="555555")
sc(ws1, 20, 4, "※年間発電量÷12",   bg=LGRAY, fg="888888")

ws1.row_dimensions[21].height = 20
sc(ws1, 21, 1, "余剰電力（参考）",      bg=LGRAY, fg="555555")
sc(ws1, 21, 2, "=MAX(0,B10/12-MIN(B9,B10/12))", bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 21, 3, "kWh/月",               bg=LGRAY, align="center", fg="555555")
sc(ws1, 21, 4, "※発電量から日中消費を引いた余剰", bg=LGRAY, fg="888888")

ws1.row_dimensions[22].height = 20
sc(ws1, 22, 1, "蓄電池自家消費（参考）", bg=LGRAY, fg="555555")
sc(ws1, 22, 2, "=MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 22, 3, "kWh/月",                bg=LGRAY, align="center", fg="555555")
sc(ws1, 22, 4, "※蓄電池容量×30日と余剰の小さい方", bg=LGRAY, fg="888888")

ws1.row_dimensions[23].height = 20
sc(ws1, 23, 1, "自家消費量（参考）",    bg=LGRAY, fg="555555")
sc(ws1, 23, 2, "=MIN(B9,B10/12)+MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 23, 3, "kWh/月",               bg=LGRAY, align="center", fg="555555")
sc(ws1, 23, 4, "※日中消費＋蓄電池消費", bg=LGRAY, fg="888888")

ws1.row_dimensions[24].height = 20
sc(ws1, 24, 1, "電気代削減額（参考）",  bg=LGRAY, fg="555555")
sc(ws1, 24, 2, "=MIN(B23*30,B8-B18)",  bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 24, 3, "円/月",                bg=LGRAY, align="center", fg="555555")
sc(ws1, 24, 4, "※自家消費量×買電単価", bg=LGRAY, fg="888888")

ws1.row_dimensions[25].height = 20
sc(ws1, 25, 1, "売電量（参考）",        bg=LGRAY, fg="555555")
sc(ws1, 25, 2, "=B10/12-MIN(B9,B10/12)-MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 25, 3, "kWh/月",               bg=LGRAY, align="center", fg="555555")
sc(ws1, 25, 4, "※発電量から自家消費を引いた売電分", bg=LGRAY, fg="888888")

ws1.row_dimensions[26].height = 20
sc(ws1, 26, 1, "売電額（参考）",        bg=LGRAY, fg="555555")
sc(ws1, 26, 2, "=ROUND(B25*24,0)",     bg=LGRAY, bold=True, align="right", fmt="#,##0", fg="555555")
sc(ws1, 26, 3, "円/月",                bg=LGRAY, align="center", fg="555555")
sc(ws1, 26, 4, "※売電量×売電単価（FIT）", bg=LGRAY, fg="888888")

sc(ws1, 27, 1, "※ 入力後「結果」シートを確認してください",
   bg=WHITE, fg="888888", merge_to="D", height=20)
ws1.cell(27, 1).font = Font(name="メイリオ", size=9, italic=True, color="888888")

# ---- 計算ステップ（縦フロー） ----
ws1.row_dimensions[28].height = 8
for col in range(1, 5):
    ws1.cell(28, col).fill = PatternFill("solid", fgColor=WHITE)

sc(ws1, 29, 1, "▼ 計算ステップ（入力値から結果までの流れ）",
   bg=GREEN, bold=True, fg=DGREEN, merge_to="D", height=20)

def step(row, label, formula, unit):
    ws1.row_dimensions[row].height = 22
    sc(ws1, row, 1, label,   bg=LGRAY, fg="333333", bold=True)
    sc(ws1, row, 2, formula, bg=LGREEN, bold=True, align="right", fmt="#,##0", fg=DGREEN)
    sc(ws1, row, 3, unit,    bg=LGRAY, align="center", fg="555555")
    sc(ws1, row, 4, "",      bg=LGRAY)

def arrow(row, text):
    ws1.row_dimensions[row].height = 13
    ws1.merge_cells(f"A{row}:D{row}")
    c = ws1.cell(row=row, column=1, value=text)
    c.font = Font(name="メイリオ", size=8, italic=True, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=3)
    c.fill = PatternFill("solid", fgColor=WHITE)
    for col in range(2, 5):
        ws1.cell(row, col).fill = PatternFill("solid", fgColor=WHITE)

step(30, "①月間発電量",      "=B10/12",                  "kWh/月")
arrow(31, "↓  MIN(自家消費量B9, 月間発電量)")
step(32, "②日中直接消費",    "=MIN(B9,B10/12)",          "kWh/月")
arrow(33, "↓  月間発電量 − 日中消費")
step(34, "③余剰電力",        "=MAX(0,B10/12-MIN(B9,B10/12))", "kWh/月")
arrow(35, "↓  MIN(蓄電池容量B7×30日, 余剰)")
step(36, "④蓄電池自家消費",  "=MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", "kWh/月")
arrow(37, "↓  日中消費 ＋ 蓄電池消費")
step(38, "⑤合計自家消費",    "=MIN(B9,B10/12)+MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", "kWh/月")
arrow(39, "↓  月間発電量 − 合計自家消費")
step(40, "⑥売電量",          "=B10/12-MIN(B9,B10/12)-MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12)))", "kWh/月")
arrow(41, "↓  MIN(合計自家消費×30円, 電気代B8 − 基本料金B18)")
step(42, "⑦電気代削減額",    "=-MIN((MIN(B9,B10/12)+MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12))))*30,B8-B18)", "円/月")
arrow(43, "↓  MAX(基本料金B18, 電気代B8 − 削減額)")
step(44, "⑧削減後電気代",    "=MAX(B18,B8-MIN((MIN(B9,B10/12)+MIN(B7*30,MAX(0,B10/12-MIN(B9,B10/12))))*30,B8-B18))", "円/月")

# =================================================================
# Sheet2: 結果シート
# =================================================================
ws2 = wb.create_sheet("結果")

for col, w in zip("ABCDEFG", [16, 18, 18, 22, 22, 26, 18]):
    ws2.column_dimensions[col].width = w
for c in "IJK":
    ws2.column_dimensions[c].width = 3

# Formula building blocks
B          = "'入力'!$B$"
mgen_      = f"({B}10/12)"                             # 月間発電量（年間発電量=B10）
daytime_   = f"MIN({B}9,{mgen_})"                     # 日中直接消費（自家消費=B9）
surplus_   = f"MAX(0,{mgen_}-{daytime_})"             # 余剰（蓄電池に充電可能分）
battery_   = f"MIN({B}7*30,{surplus_})"               # 蓄電池自家消費
self_      = f"({daytime_}+{battery_})"               # 合計自家消費
sell_      = f"({mgen_}-{self_})"                     # 売電量（必ず0以上）
save_      = f"MIN({self_}*{B}14,{B}8-{B}18)"        # 電気代削減額（最大=電気代−基本料金）
red_       = f"MAX({B}18,{B}8-{save_})"               # 削減後電気代（基本料金フロア）

def cum_sell(yr):
    if yr <= 4:
        return f"ROUND({sell_}*{B}15*{yr*12},0)"
    elif yr <= 10:
        return f"ROUND({sell_}*{B}15*48,0)+ROUND({sell_}*{B}16*{(yr-4)*12},0)"
    else:
        return f"ROUND({sell_}*{B}15*48,0)+ROUND({sell_}*{B}16*72,0)+ROUND({sell_}*{B}17*{(yr-10)*12},0)"

def sell_yr(yr):
    if yr <= 4:
        return f"=ROUND({sell_}*{B}15*12,0)"
    elif yr <= 10:
        return f"=ROUND({sell_}*{B}16*12,0)"
    else:
        return f"=ROUND({sell_}*{B}17*12,0)"

# ---- Row 1: Logo ----
sc(ws2, 1, 1, "", bg=WHITE, merge_to="F", height=80)
add_logo_centered(ws2, row_height_px=80, total_width_px=800, logo_path=LOGO, logo_h_px=64)

# ---- Row 2: Title ----
ws2.row_dimensions[2].height = 36
ws2.merge_cells("A2:C2")
cc = ws2.cell(row=2, column=1, value="='入力'!B5")
cc.font = Font(name="メイリオ", size=22, bold=True, color="222222")
cc.alignment = Alignment(horizontal="right", vertical="center")
cc.fill = PatternFill("solid", fgColor=WHITE)
cc.border = BORDER

ws2.merge_cells("D2:F2")
tc = ws2.cell(row=2, column=4, value="様邸　光熱費シュミレーション")
tc.font = Font(name="メイリオ", size=18, bold=True, color="333333")
tc.alignment = Alignment(horizontal="left", vertical="center")
tc.fill = PatternFill("solid", fgColor=WHITE)
tc.border = BORDER

# ---- Row 3: Specs horizontal (A:F, 3ペア) ----
ws2.row_dimensions[3].height = 24
spec_pairs = [
    ("パネル容量",  f"='入力'!B6&\" kW\"",   1, 2),
    ("蓄電池容量",  f"='入力'!B7&\" kWh\"",  3, 4),
    ("年間発電量",  f"='入力'!B10&\" kWh\"", 5, 6),
]
for lbl, val, lc_col, vc_col in spec_pairs:
    lc = ws2.cell(row=3, column=lc_col, value=lbl)
    lc.fill = PatternFill("solid", fgColor=GREEN)
    lc.font = Font(name="メイリオ", size=10, bold=True, color=DGREEN)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = BORDER
    vc = ws2.cell(row=3, column=vc_col, value=val)
    vc.fill = PatternFill("solid", fgColor=LGREEN)
    vc.font = Font(name="メイリオ", size=11, bold=True, color=DGREEN)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = BORDER

# ---- Row 4: spacer ----
ws2.row_dimensions[4].height = 6
fill_row(ws2, 4, WHITE)

# ---- Rows 16-34: chart overlay area ----
for r in range(16, 35):
    ws2.row_dimensions[r].height = 15
    fill_row(ws2, r, "FAFAFA")

# ---- Chart data in hidden cols I(9) J(10) K(11), rows 2-12 ----
# White font on white background = visually invisible
def hidden_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=WHITE)
    c.font = Font(name="メイリオ", size=9, color=WHITE)
    c.border = Border()

hidden_cell(ws2, 2, 9,  "年")
hidden_cell(ws2, 2, 10, "現在の電気代（累積）")
hidden_cell(ws2, 2, 11, "実質支払額（累積）")
for yr in range(1, 11):
    dr = yr + 2
    hidden_cell(ws2, dr, 9,  yr)
    hidden_cell(ws2, dr, 10, f"={B}8*12*{yr}")
    hidden_cell(ws2, dr, 11, f"={red_}*12*{yr}-({cum_sell(yr)})")

# ---- BarChart ----
chart = BarChart()
chart.type     = "col"
chart.grouping = "clustered"
chart.title    = None
chart.style    = 2
chart.y_axis.numFmt = '#,##0'

data_ref = Reference(ws2, min_col=10, max_col=11, min_row=2, max_row=12)
cats_ref = Reference(ws2, min_col=9,  min_row=3, max_row=12)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

chart.series[0].graphicalProperties.solidFill = "1F4E79"   # 設備なし累計: ネイビー
chart.series[1].graphicalProperties.solidFill = "E8730A"   # 実質支払額: オレンジ
chart.series[1].invertIfNegative = False  # 負の値でも色反転させない

for s in chart.series:
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.showSerName = False
    s.dLbls.showCatName = False
    s.dLbls.showLegendKey = False
    s.dLbls.numFmt = '#,##0'

# 横幅をA-G列の合計幅に合わせる (20+13+13+16+16+18+18=114chars × 7px / 96dpi × 2.54 ≈ 21cm)
chart.width  = 23
chart.height = 10
chart.legend.position = "b"
ws2.add_chart(chart, "A16")

# ---- Row 5: Summary section header ----
sc(ws2, 5, 1, "▼ シミュレーション結果",
   bg=GREEN, bold=True, fg=DGREEN, merge_to="F", height=22)

# ---- Rows 6-7: 30年間の実質削減額 全幅バナー（ヘッダー直下）----
jitsu_30 = f"={save_}*12*30+{cum_sell(30)}"
ws2.merge_cells("A6:F6")
bh = ws2.cell(row=6, column=1, value="30年間の実質削減額は")
bh.fill = PatternFill("solid", fgColor=RED)
bh.font = Font(name="メイリオ", size=13, bold=True, color=WHITE)
bh.alignment = Alignment(horizontal="center", vertical="center")
bh.border = BORDER
ws2.row_dimensions[6].height = 26

ws2.merge_cells("A7:F7")
bv = ws2.cell(row=7, column=1, value=jitsu_30)
bv.fill = PatternFill("solid", fgColor=LRED)
bv.font = Font(name="メイリオ", size=26, bold=True, color=RED)
bv.alignment = Alignment(horizontal="center", vertical="center")
bv.number_format = '#,##0"円"'
bv.border = RBORDER
ws2.row_dimensions[7].height = 48

# ---- Row 8: Column headers ----
ws2.row_dimensions[8].height = 28
col_hdrs = [("", LGRAY), ("1カ月", GREEN), ("4年目（累計）", GREEN),
            ("10年目（累計）", GREEN), ("20年目（累計）", GREEN), ("30年目（累計）", GREEN)]
for col, (lbl, bg_c) in enumerate(col_hdrs, 1):
    c = sc(ws2, 8, col, lbl, bg=bg_c, bold=True,
           fg=DGREEN if bg_c == GREEN else "555555", align="center", wrap=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---- Rows 9-12: Summary data ----
summary_rows = [
    ("現在の電気代（太陽光なし）",
     f"={B}8", f"={B}8*12*4", f"={B}8*12*10", f"={B}8*12*20", f"={B}8*12*30",
     LGRAY, "333333"),
    ("電気代削減額",
     f"={save_}", f"={save_}*12*4", f"={save_}*12*10", f"={save_}*12*20", f"={save_}*12*30",
     WHITE, "333333"),
    ("削減後電気代",
     f"={red_}", f"={red_}*12*4", f"={red_}*12*10", f"={red_}*12*20", f"={red_}*12*30",
     WHITE, "333333"),
    ("売電収入",
     f"=ROUND({sell_}*{B}15,0)",
     f"={cum_sell(4)}", f"={cum_sell(10)}", f"={cum_sell(20)}", f"={cum_sell(30)}",
     WHITE, "333333"),
]
for i, (label, v1, v2, v3, v4, v5, bg, fg_c) in enumerate(summary_rows):
    row = 9 + i
    ws2.row_dimensions[row].height = 26
    sc(ws2, row, 1, label, bg=LGRAY, fg=fg_c)
    for col, val in enumerate([v1, v2, v3, v4, v5], 2):
        sc(ws2, row, col, val, bg=bg, align="center", fg=fg_c, fmt='#,##0')

# ---- Row 13: spacer ----
ws2.row_dimensions[13].height = 6
fill_row(ws2, 13, WHITE)

# ---- Row 14: 実質削減額 (RED) ----
ws2.row_dimensions[14].height = 34
sc(ws2, 14, 1, "実質削減額", bg=RED, bold=True, fg=WHITE, size=12, align="center")
jitsu_vals = [
    f"={save_}+ROUND({sell_}*{B}15,0)",
    f"={save_}*12*4+{cum_sell(4)}",
    f"={save_}*12*10+{cum_sell(10)}",
    f"={save_}*12*20+{cum_sell(20)}",
    f"={save_}*12*30+{cum_sell(30)}",
]
for col, val in enumerate(jitsu_vals, 2):
    sc(ws2, 14, col, val, bg=LRED, bold=True, align="center", fg=RED, size=12, fmt='#,##0')

# ---- Row 15: spacer ----
ws2.row_dimensions[15].height = 8
fill_row(ws2, 15, WHITE)

# ---- Row 35: spacer ----
ws2.row_dimensions[35].height = 8
fill_row(ws2, 35, WHITE)

# ---- Row 36: 年別詳細 header ----
sc(ws2, 36, 1, "▼ 年別詳細シミュレーション（1〜10年目・20年目・30年目）",
   bg=GREEN, bold=True, fg=DGREEN, merge_to="F", height=22)

# ---- Row 37: year table headers ----
hdrs = ["年目", "削減後の電気代\n(累積)", "売電収入\n(累積)",
        "設備なしの電気代\n(累積)", "削減累積", "実質支払額\n(累積)"]
for col, h in enumerate(hdrs, 1):
    c = sc(ws2, 37, col, h, bg=GREEN, bold=True, fg=DGREEN, align="center", height=32, wrap=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---- Rows 38-49: year data (1-10, 20, 30) — all columns cumulative ----
years_to_show = list(range(1, 11)) + [20, 30]
for idx, yr in enumerate(years_to_show):
    row = 38 + idx
    ws2.row_dimensions[row].height = 20
    alt_bg = LGREEN if idx % 2 == 0 else WHITE  # LGREEN=薄黄
    cs = cum_sell(yr)

    row_vals = [
        yr,
        f"={red_}*12*{yr}",          # 削減後電気代 累積
        f"={cs}",                     # 売電収入 累積
        f"={B}8*12*{yr}",             # 設備なし電気代 累積
        f"={save_}*12*{yr}+{cs}",     # 削減累積
        f"={red_}*12*{yr}-({cs})",    # 実質支払額 累積
    ]
    fmts = [None, '#,##0"円"', '#,##0"円"', '#,##0"円"', '#,##0"円"', '#,##0"円"']

    for col, (val, fmt) in enumerate(zip(row_vals, fmts), 1):
        is_total = col == 5
        is_jitsu = col == 6
        fg_c = RED if is_jitsu else (DGREEN if is_total else ("555555" if col == 1 else "333333"))
        bold_c = is_total or is_jitsu
        sc(ws2, row, col, val, bg=alt_bg, align="center", fmt=fmt,
           fg=fg_c, bold=bold_c)

# ---- Notes ----
ws2.row_dimensions[51].height = 8
sc(ws2, 52, 1, "※ 売電単価：年1〜4年目=24円 / 年5〜10年目=8.3円 / 年11〜30年目=7円",
   bg=WHITE, merge_to="F", height=18)
ws2.cell(52, 1).font = Font(name="メイリオ", size=9, italic=True, color="888888")
sc(ws2, 53, 1, "※ 本シミュレーションは発電量・削減額を保証するものではありません。",
   bg=WHITE, merge_to="F", height=18)
ws2.cell(53, 1).font = Font(name="メイリオ", size=9, italic=True, color="888888")

# =================================================================
# 印刷設定 — 入力シート
# =================================================================
ws1.page_setup.orientation  = 'portrait'
ws1.page_setup.paperSize    = 9          # A4
ws1.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws1.page_setup.fitToWidth   = 1
ws1.page_setup.fitToHeight  = 1
ws1.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                               header=0.2, footer=0.2)
ws1.print_area = 'A1:D19'
ws1.print_options.gridLines = False

# =================================================================
# 印刷設定 — 結果シート（縦A4 × 2ページ）
#   P.1 : ロゴ＋タイトル＋グラフ＋サマリー（rows 1-34）
#   P.2 : 年別詳細テーブル（rows 35-53）
# =================================================================
ws2.page_setup.orientation  = 'portrait'
ws2.page_setup.paperSize    = 9
ws2.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws2.page_setup.fitToWidth   = 1
ws2.page_setup.fitToHeight  = 0          # ページ数は自動
ws2.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                               header=0.2, footer=0.2)
ws2.print_area = 'A1:F53'
ws2.print_options.gridLines = False

# サマリーと年別詳細の間で改ページ（row34の後）
ws2.row_breaks.append(Break(id=35))

# =================================================================
# Sheet3: 組み合わせマトリクス（旧ロジックで売電がマイナスになる組み合わせ）
# =================================================================
ws3 = wb.create_sheet("組み合わせ確認")

PANEL_KWS   = [3, 4, 5, 6, 7, 8, 9, 10]
BATTERY_KWHS = [0, 3, 5, 7, 9.5, 10, 12, 15, 16]
SELF_CONSUME = 100  # 固定

RED_CELL  = "FFCCCC"
GRN_CELL  = "CCEECC"
YLW_CELL  = "FFF9C4"

ws3.column_dimensions['A'].width = 18
for col_idx in range(2, 2 + len(BATTERY_KWHS)):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 12

# タイトル
t = ws3.cell(row=1, column=1, value="売電量マトリクス（旧ロジック：自家消費=100+蓄電池×30）")
t.font = Font(name="メイリオ", bold=True, size=11)
ws3.merge_cells(f"A1:{get_column_letter(1+len(BATTERY_KWHS))}1")
ws3.row_dimensions[1].height = 24

note = ws3.cell(row=2, column=1, value="※ 赤セル = 売電量がマイナス（発電量 < 自家消費） / 緑セル = 売電あり")
note.font = Font(name="メイリオ", size=9, color="666666", italic=True)
ws3.merge_cells(f"A2:{get_column_letter(1+len(BATTERY_KWHS))}2")
ws3.row_dimensions[2].height = 18

# ヘッダー行：蓄電池容量
h = ws3.cell(row=3, column=1, value="パネル容量 ↓ / 蓄電池 →")
h.font = Font(name="メイリオ", bold=True, size=9, color=DGREEN)
h.fill = PatternFill("solid", fgColor=GREEN)
h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
h.border = BORDER
ws3.row_dimensions[3].height = 30

for j, bat in enumerate(BATTERY_KWHS):
    c = ws3.cell(row=3, column=2+j, value=f"{bat} kWh" if bat > 0 else "なし")
    c.font = Font(name="メイリオ", bold=True, size=9, color=DGREEN)
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

# データ行
for i, panel in enumerate(PANEL_KWS):
    row = 4 + i
    ws3.row_dimensions[row].height = 22
    monthly_gen = panel * 1100 / 12

    # 行ヘッダー
    rh = ws3.cell(row=row, column=1, value=f"{panel} kW（{panel*1100:,} kWh/年）")
    rh.font = Font(name="メイリオ", bold=True, size=9, color=DGREEN)
    rh.fill = PatternFill("solid", fgColor=GREEN)
    rh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh.border = BORDER

    for j, bat in enumerate(BATTERY_KWHS):
        old_self = SELF_CONSUME + bat * 30
        old_sell = monthly_gen - old_self
        rounded  = round(old_sell)

        c = ws3.cell(row=row, column=2+j)
        c.value = rounded
        c.number_format = '#,##0 "kWh"'
        c.font = Font(name="メイリオ", size=9, bold=(rounded < 0))
        c.fill = PatternFill("solid", fgColor=RED_CELL if rounded < 0 else GRN_CELL)
        c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c.border = BORDER

# =================================================================
out = "/Users/iwa/claude/solar-simulator/太陽光シミュレーションエクセル（20260422エンパワー修正）.xlsx"
wb.save(out)
print(f"saved: {out}")
